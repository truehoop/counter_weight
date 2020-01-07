from openpyxl import load_workbook

workbook = load_workbook(filename = 'eval_result.xlsx')
print(workbook.sheetnames)
sheet = workbook.active

read_mode = 'TYPE'

sheet_ranges = workbook['ID랑title기준 동일기사삭제']
sheet_ids = sheet_ranges['A']
result = {}

if read_mode == 'PN':
    sheet_values = sheet_ranges['J']
    for idx in range(len(sheet_ids)):
        sheet_val = sheet_values[idx].value
        if sheet_val is None:
            continue
        elif sheet_val == 'P' or sheet_val == 'p':
            result[sheet_ids[idx].value] = 0
        elif sheet_val == 'N' or sheet_val == 'n':
            result[sheet_ids[idx].value] = 1
        else:
            continue

elif read_mode == 'TYPE':
    sheet_values = sheet_ranges['G']
    for idx in range(len(sheet_ids)):
        if idx == 0:
            continue
        sheet_val = sheet_values[idx].value
        if sheet_val is None or sheet_val == '?':
            continue
        elif sheet_val == 'X':
            result[sheet_ids[idx].value] = 17
            continue
        result[sheet_ids[idx].value] = int(sheet_val)-1

def get_sheet_value():
    return result

print('sheet_ids', len(sheet_ids))
# print(result)
