# -*- coding:utf-8 -*-

from openpyxl import load_workbook, Workbook

# 1. 키워드 입력

workbook = load_workbook(filename = 'input_keyword.xlsx')
sheet = workbook.active

sheet_ranges = workbook['시트 1']
keyword1 = sheet_ranges['A']
keyword2 = sheet_ranges['B']

print(keyword1)
print(keyword2)

keyword1 = tuple(x for x in keyword1 if (x.value is not None and x.value != '키워드1'))
keyword2 = tuple(x for x in keyword2 if (x.value is not None and x.value != '키워드2'))


# jiname.xlsx
jiname_wb = load_workbook(filename = 'jiname.xlsx')
sheet_ranges = jiname_wb['Sheet1']
jiname = sheet_ranges['A']
jiname = tuple(x.value for x in jiname if (x.value is not None and x.value != '시도명칭'))

# ext_key.xlsx
ext_key_wb = load_workbook(filename = 'ext_key.xlsx')
sheet_ranges = ext_key_wb['Sheet1']
ext_key = sheet_ranges['A']
ext_key = tuple(x.value for x in ext_key if (x.value is not None and x.value != 'public'))

# big_com.xlsx
big_com_wb = load_workbook(filename = 'big_com.xlsx')
sheet_ranges = big_com_wb['Sheet1']
big_com = sheet_ranges['A']
big_com = tuple(x.value for x in big_com if (x.value is not None and x.value != 'big_com'))



# 2. 엑셀에서 기관명 추출

def comprehension(a, b):
    return [x for x in a if x not in b]

def set_approach(a,b):
    return list(set(a)-set(b))

wb_result = Workbook()

for keywordA in keyword1:
    for keywordB in keyword2:
        keyword_combination = keywordA.value +'_'+ keywordB.value
        keyword_wb = load_workbook(filename = 'value_news/result' + keyword_combination + '.xlsx')

        sheet_ranges = keyword_wb['sheet']
        org_names = sheet_ranges['M']

        org_names = []

        ws = keyword_wb.active
        for row in ws.iter_rows(min_row=2, min_col=13, max_col=13):
            for cell in row:
                splited_value = cell.value.split(",")
                org_names = org_names + splited_value

        org_names = set(org_names)
        print(keyword_combination)
        print(org_names)

        # 3. 구분한 기관명 중 내가 아는 기관명 제외
        print(len(org_names))
        org_names = set_approach(org_names, jiname)
        org_names = set_approach(org_names, ext_key)
        org_names = set_approach(org_names, big_com)
        print(len(org_names))
        ws = wb_result.create_sheet(title=keyword_combination)
        for org_name in org_names:
            ws.append([org_name])


# print(jiname)
# print(ext_key)
# print(big_com)

# 4. 저장
from datetime import datetime
current_time = datetime.now().strftime("%H-%M-%S")
file_name = '기관명_키워드' + current_time + '.xlsx'
print(file_name)
wb_result.save(filename = file_name)








