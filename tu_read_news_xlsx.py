# -*- coding:utf-8 -*-

# 1. 키워드 입력
from openpyxl import load_workbook

workbook = load_workbook(filename = 'input_keyword.xlsx')
print(workbook.sheetnames)
sheet = workbook.active

sheet_ranges = workbook['시트 1']
keyword1 = sheet_ranges['A']
keyword2 = sheet_ranges['B']
# worksheet.iter_rows(min_row=5) 사용하면 반복가능

result = {}
print(sheet_ranges)
print(keyword1)
print(keyword2)

keyword1 = tuple(x for x in keyword1 if (x.value is not None and x.value != '키워드1'))
keyword2 = tuple(x for x in keyword2 if (x.value is not None and x.value != '키워드2'))

#
# for keyword in keyword1:
#     print(1, keyword.value)
#     if keyword.value is None:
#         keyword1.remove(keyword)
#
# for keyword in keyword2:
#     if keyword.value == '':
#         keyword1.remove(keyword)


print(keyword1)
print(keyword2)



# 2. 뉴스 엑셀 다운로드
import requests
import time

url = 'https://www.bigkinds.or.kr/api/news/download.do'

new_headers = {
"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.135 Safari/537.36",
"Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9",
"Cookie": "Bigkinds=86D85B26F7492533CE246A661673C532",
"Refer": "https://www.bigkinds.or.kr/v2/news/search.do",
"Content-Type": "application/x-www-form-urlencoded"
}

payload = {
    "jsonSearchParam": '{"indexName":"news","searchKey":"비건 샴푸","searchKeys":[{}],"byLine":"","searchFilterType":"1","searchScopeType":"1","searchSortType":"date","sortMethod":"date","mainTodayPersonYn":"","startDate":"2020-01-01","endDate":"2021-03-17","newsIds":[],"categoryCodes":[],"providerCodes":[],"incidentCodes":[],"networkNodeType":"","topicOrigin":"","dateCodes":[],"editorialIs":false,"startNo":1,"resultNumber":10,"isTmUsable":false,"isNotTmUsable":false,"sectionDiv":"1000","realURI":"/api/news/previewData.do"}'
}

session = requests.Session()

for keywordA in keyword1:
    for keywordB in keyword2:
        payload = {
            "jsonSearchParam": '{"indexName":"news","searchKey":"'
                               + keywordA.value + ' ' + keywordB.value
                               + '","searchKeys":[{}],"byLine":"","searchFilterType":"1","searchScopeType":"1","searchSortType":"date","sortMethod":"date","mainTodayPersonYn":"","startDate":"2020-01-01","endDate":"2021-03-17","newsIds":[],"categoryCodes":[],"providerCodes":[],"incidentCodes":[],"networkNodeType":"","topicOrigin":"","dateCodes":[],"editorialIs":false,"startNo":1,"resultNumber":10,"isTmUsable":false,"isNotTmUsable":false,"sectionDiv":"1000","realURI":"/api/news/previewData.do"}'
        }
        print(payload)
        r = session.post(url, headers=new_headers, data=payload)
        open('value_news/result'+ keywordA.value + '_' + keywordB.value +'.xlsx', 'wb').write(r.content)
        time.sleep(5)






